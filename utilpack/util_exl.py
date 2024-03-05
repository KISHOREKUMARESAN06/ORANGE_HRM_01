import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

class XLU:

    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    def getCoulmnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_coulmn)

    def readData(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(rownum, columnno).value

    def writeData(file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(rownum, columnno).value = data
        workbook.save(file)

    def fillGreenColor(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        greenFill = PatternFill(start_color='60b212',
                                end_color='60b212',
                                fill_type='solid')
        sheet.cell(rownum, columnno).fill = greenFill
        workbook.save(file)

    def fillRedColor(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        redFill = PatternFill(start_color='ff0000',
                              end_color='ff0000',
                              fill_type='solid')
        sheet.cell(rownum, columnno).fill = redFill
        workbook.save(file)


    ## printing expected result functions when all the performance successfully.

    def login_exp_res(self):
        print(XLU.readData(exl_file_path, sheet1, 2, 8))

    def invalid_login_exp_res(self):
        print(XLU.readData(exl_file_path, sheet1, 3, 8))

    def pim01_exp_res(self):
        print(XLU.readData(exl_file_path, sheet1, 4, 8))

    def pim02_exp_res(self):
        print(XLU.readData(exl_file_path, sheet1, 5, 8))

    def pim03_exp_res(self):
        print(XLU.readData(exl_file_path, sheet1, 6, 8))


    ## date_and_time function for print the current date,time in to the excel sheet
    def date_and_time(self):

        row = XLU.getRowCount(exl_file_path, sheet1)

        for r in range(2 , row+1):
            # adding the current date to excel
            run_date = datetime.now().strftime('%Y-%m-%d')
            XLU.writeData(exl_file_path, sheet1, r, 11, run_date)

            # adding the current time to excel
            run_time = datetime.now().time().strftime('%H:%M:%S')
            XLU.writeData(exl_file_path, sheet1, r, 12, run_time)

    # get the data from the Excel file for adding operation.
    def adding_datas(self, exl_file_path, sheet2):

        added= []
        f_name = XLU.readData(exl_file_path, sheet2, 4,4)
        m_name = XLU.readData(exl_file_path , sheet2, 4, 5)
        l_name = XLU.readData(exl_file_path , sheet2, 4, 6)
        emp_id = XLU.readData(exl_file_path , sheet2, 4, 7)
        u_name = XLU.readData(exl_file_path , sheet2, 4, 8 )
        pwd = XLU.readData(exl_file_path , sheet2, 4, 9)
        oth_id = XLU.readData(exl_file_path , sheet2, 4, 10)
        dri_no = XLU.readData(exl_file_path , sheet2, 4, 11)
        l_ex_dt = XLU.readData(exl_file_path , sheet2, 4, 12)
        dob = XLU.readData(exl_file_path , sheet2, 4, 13)
        gen = XLU.readData(exl_file_path , sheet2, 4, 14)
        T_field = XLU.readData(exl_file_path , sheet2, 4, 15)

        add = (f_name,m_name,l_name,emp_id,u_name,pwd,oth_id,dri_no,l_ex_dt,dob,gen,T_field)
        add.append(added)
        return added

    # get the data from the Excel file for editing operation.
    def editing_datas(self,exl_file_path, sheet2):

        edited = []

        m_name = XLU.readData(exl_file_path, sheet2, 5, 5)
        l_name = XLU.readData(exl_file_path, sheet2, 5, 6)
        emp_id = XLU.readData(exl_file_path, sheet2, 5, 7)

        edit = (m_name, l_name, emp_id)
        edit.append(edited)
        return edited

